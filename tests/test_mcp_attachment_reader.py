import io

from openpyxl import Workbook

from mcp_email_server.attachment_reader import extract_attachment_text


def test_extract_attachment_text_csv():
    payload = b"col1,col2\nalpha,beta\n"
    text = extract_attachment_text("sample.csv", payload)
    assert "col1 | col2" in text
    assert "alpha | beta" in text


def test_extract_attachment_text_xlsx():
    wb = Workbook()
    ws = wb.active
    ws.title = "SheetA"
    ws["A1"] = "sku"
    ws["B1"] = "price"
    ws["A2"] = "SKU-1"
    ws["B2"] = 12.5
    buf = io.BytesIO()
    wb.save(buf)
    text = extract_attachment_text("sample.xlsx", buf.getvalue())
    assert "Sheet: SheetA" in text
    assert "SKU-1" in text
