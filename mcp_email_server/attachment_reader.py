from __future__ import annotations

import csv
import io
from pathlib import Path

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader


def extract_attachment_text(filename: str, payload: bytes) -> str:
    suffix = Path(filename).suffix.lower()
    if suffix == ".pdf":
        return _read_pdf(payload)
    if suffix == ".docx":
        return _read_docx(payload)
    if suffix == ".csv":
        return _read_csv(payload)
    if suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        return _read_xlsx(payload)
    return ""


def _read_pdf(payload: bytes) -> str:
    reader = PdfReader(io.BytesIO(payload))
    chunks: list[str] = []
    for page in reader.pages[:20]:
        chunks.append(page.extract_text() or "")
    return _normalize(" ".join(chunks))


def _read_docx(payload: bytes) -> str:
    doc = Document(io.BytesIO(payload))
    text = " ".join(p.text for p in doc.paragraphs if p.text.strip())
    return _normalize(text)


def _read_csv(payload: bytes) -> str:
    data = payload.decode("utf-8", errors="ignore")
    reader = csv.reader(io.StringIO(data))
    rows: list[str] = []
    for idx, row in enumerate(reader):
        rows.append(" | ".join(col.strip() for col in row))
        if idx >= 100:
            break
    return _normalize("\n".join(rows))


def _read_xlsx(payload: bytes) -> str:
    wb = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    chunks: list[str] = []
    for ws in wb.worksheets[:3]:
        chunks.append(f"[Sheet: {ws.title}]")
        for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True)):
            cells = [str(v) for v in row if v is not None and str(v).strip()]
            if cells:
                chunks.append(" | ".join(cells))
            if r_idx >= 50:
                break
    return _normalize("\n".join(chunks))


def _normalize(text: str) -> str:
    return " ".join(text.split())[:6000]
